import pandas as pd
from typing import List, Dict, Optional, Tuple, Any
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import sys
from loguru import logger

from src.services.product_config_service import ProductConfigService
from src.services.platform_config_service import ColumnMappingService
from src.services.store_address_service import StoreAddressService
from src.services.report_adapters import (
    ShoplineAdapter, MixxAdapter, C2CAdapter, AoshiAdapter,
    BaseAdapter, StandardOrderItem, ConversionResult
)

class UnifiedOrderProcessor:
    def __init__(self, product_service: ProductConfigService):
        self.product_service = product_service

    def process(self, items: List[StandardOrderItem]) -> List[Dict]:
        """
        Process standard items into final shipping rows with box calculation.
        """
        # Group by order_id
        orders: Dict[str, List[StandardOrderItem]] = {}
        for item in items:
            if item.order_id not in orders:
                orders[item.order_id] = []
            orders[item.order_id].append(item)

        final_rows = []

        for order_id, order_items in orders.items():
            # Create rows for items
            current_order_rows = []
            for item in order_items:
                row = self._create_output_row(item)
                current_order_rows.append(row)
            
            # Calculate box
            box_info = self._calculate_box(order_items)
            
            # Create box row
            if current_order_rows:
                # Use first item as template for box row common fields
                box_row = current_order_rows[0].copy()
                box_row.update({
                    "商品編號": box_info["code"],
                    "商品名稱": box_info["name"],
                    "訂購數量": "1",
                    "品項備註": "箱子"
                })
                # Add box row to the end
                current_order_rows.append(box_row)
            
            final_rows.extend(current_order_rows)

        return final_rows

    def _create_output_row(self, item: StandardOrderItem) -> Dict:
        return {
            "貨主編號": "A442",
            "貨主單號\n(不同客戶端、不同溫層要分單)": item.order_id,
            "客戶端代號(店號)": item.receiver_name,
            "訂購日期": item.order_date,
            "商品編號": item.product_code,
            "商品名稱": item.product_name,
            "訂購數量": str(item.quantity),
            "配送方式\nFT : 逢泰配送\nTcat : 黑貓宅配\n全家到府取貨": item.delivery_method,
            "收貨人姓名": item.receiver_name,
            "收貨人地址": item.receiver_address,
            "收貨人聯絡電話": item.receiver_phone,
            "訂單 / 宅配單備註": item.original_row.get("order_mark_fmt", item.order_mark), # Can format here if needed
            "指定配送溫層\n001：常溫\n002：冷藏\n003：冷凍": "003",
            "品項備註": "",
            "到貨時段\n1: 13點前\n2: 14~18\n3: 不限時": item.arrival_time
        }

    def _calculate_box(self, items: List[StandardOrderItem]) -> Dict:
        grand_total = 0
        for item in items:
            if not item.product_code:
                continue
            
            # Get qty per pack from config
            qty_per_pack = self.product_service.get_product_qty(item.product_code)
            grand_total += qty_per_pack * item.quantity
        
        if grand_total <= 14:
            return {"code": "box60-EA", "name": "60公分紙箱"}
        elif grand_total <= 47:
            return {"code": "box90-EA", "name": "90公分紙箱"}
        else:
            return {"code": "ERROR-需拆單", "name": "ERROR-需拆單"}

class ReportService:
    def __init__(self):
        self.product_service = ProductConfigService()
        self.config_service = ColumnMappingService()
        self.store_address_service = StoreAddressService()
        self.processor = UnifiedOrderProcessor(self.product_service)

    def _get_adapter(self, platform: str) -> BaseAdapter:
        """Get adapter for specific platform processing logic."""
        platform = platform.lower()
        if platform == "shopline":
            return ShoplineAdapter(self.product_service, self.config_service)
        elif platform == "mixx":
            return MixxAdapter(self.product_service, self.config_service)
        elif platform == "c2c":
            return C2CAdapter(self.product_service, self.config_service)
        elif platform == "aoshi":
            return AoshiAdapter(self.product_service, self.config_service)
        else:
            # Default to Shopline adapter for general processing
            logger.warning(f"Unknown platform '{platform}', using Shopline adapter")
            return ShoplineAdapter(self.product_service, self.config_service)

    def generate_report(self, file_content: bytes, filename: str, platform: str = "shopline") -> Tuple[BytesIO, Dict]:
        """
        Generate Excel report from input file.
        Platform parameter determines processing logic (date parsing, product lookup, etc.)
        Column mapping is unified across all platforms.
        Returns (Excel Bytes, Summary Dict)
        """
        import time
        start_time = time.time()
        logger.info("[Perf] Starting report generation...")

        # Load config
        t0 = time.time()
        self.product_service.load_config()
        self.config_service.load_config()
        logger.info(f"[Perf] Config loaded in {time.time() - t0:.4f}s")

        # Read input file
        try:
            t0 = time.time()
            df = pd.read_excel(BytesIO(file_content))
            logger.info(f"[Perf] Excel read in {time.time() - t0:.4f}s (Rows: {len(df)})")
        except Exception as e:
            raise ValueError(f"Failed to read Excel file: {e}")

        # Validate that we can find required columns
        validation = self.config_service.validate_columns(list(df.columns))
        missing = [f for f, found in validation.items() if not found]
        if missing:
            logger.warning(f"Some fields not found in Excel: {missing}")

        # Get adapter (platform determines processing logic, not column mapping)
        adapter = self._get_adapter(platform)
        logger.info(f"[Perf] Using adapter: {platform}")
        
        # Convert to standard items
        t0 = time.time()
        result = adapter.convert(df, self.store_address_service)
        logger.info(f"[Perf] Adapter conversion finished in {time.time() - t0:.4f}s")
        
        # Process items (Box calc, formatting)
        t0 = time.time()
        final_rows = self.processor.process(result.items)
        logger.info(f"[Perf] Processor finished in {time.time() - t0:.4f}s")
        
        # Generate Output Excel
        t0 = time.time()
        output_buffer = self._create_excel(final_rows)
        logger.info(f"[Perf] Excel generation finished in {time.time() - t0:.4f}s")
        
        total_time = time.time() - start_time
        logger.info(f"[Perf] Total execution time: {total_time:.4f}s")

        summary = {
            "total_orders": len(set(item.order_id for item in result.items)),
            "total_rows": len(final_rows),
            "errors": result.errors,
            "platform": platform
        }

        # Print detailed stats to console for verification
        print("\n" + "="*50)
        print(f" Report Generation Summary ({platform})")
        print("="*50)
        print(f" Total Orders Processed: {summary['total_orders']}")
        print(f" Total Output Rows:      {summary['total_rows']}")
        print(f" Errors Found:           {len(summary['errors'])}")
        print(f" Time Taken:             {total_time:.2f}s")
        if summary['errors']:
            print("-" * 30)
            for err in summary['errors'][:5]:
                print(f" * [{err['severity'].upper()}] Order {err['order_id']}: {err['message']}")
            if len(summary['errors']) > 5:
                print(f" ... and {len(summary['errors']) - 5} more errors.")
        print("="*50 + "\n")
        
        return output_buffer, summary

    def _create_excel(self, rows: List[Dict]) -> BytesIO:
        template_path = "src/assets/report_template.xlsx"
        
        # User requested to remove fallback - Strict template usage
        try:
            wb = load_workbook(template_path)
            ws = wb.active
            print(f"Loaded template from {template_path}")
        except FileNotFoundError:
            raise FileNotFoundError(f"Template file missing at {template_path}. Please ensure the template exists.")

        if not rows:
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        # Get headers from template (row 1)
        # Verify valid headers exist
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)
        
        # If template is empty/new, use keys from first row
        if not headers and rows:
            headers = list(rows[0].keys())
            ws.append(headers)

        # Ensure freeze panes (A-H locked)
        ws.freeze_panes = "I2"

        # Data
        # Start writing from row 2
        start_row = 2
        
        for i, row_data in enumerate(rows):
            row_num = start_row + i
            for col_idx, header in enumerate(headers, 1):
                # Get value by header name
                val = row_data.get(header, "")
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                
                # Apply legacy-style formatting per cell
                self._apply_cell_format(cell)

        # Auto-adjust dimensions like legacy
        self._auto_adjust_dimensions(ws)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _apply_cell_format(self, cell):
        val_str = str(cell.value) if cell.value is not None else ""
        is_error = "ERROR" in val_str
        is_nan = val_str.lower() == "nan"

        # Legacy font style
        cell.font = Font(name="微軟正黑體", size=11, bold=is_nan, color="FF0000" if (is_error or is_nan) else "000000")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _auto_adjust_dimensions(self, sheet):
        # Helper to estimate visual width
        def get_visual_width(s):
            if not s: return 0
            w = 0
            for char in str(s):
                # Rough estimate: Wide characters (Chinese, etc) count as 1.7-2.0
                if ord(char) > 127:
                    w += 2.0
                else:
                    w += 1.1 # Reduced from 1.2 to standard
            return w

        # Column width
        for column_cells in sheet.columns:
            max_width = 0
            col_letter = column_cells[0].column_letter
            
            # Check all cells including header
            for cell in column_cells:
                try:
                    val = cell.value
                    if val:
                        # Handle multiline - take max of lines
                        lines = str(val).split('\n')
                        for line in lines:
                            w = get_visual_width(line)
                            if w > max_width:
                                max_width = w
                except:
                    pass
            
            # Heuristic: Min width 12, Max 60, Padding +2
            adjusted_width = min(max(max_width + 2, 12), 60)
            sheet.column_dimensions[col_letter].width = adjusted_width

        # Row height
        for row in sheet.rows:
            max_height = 0
            for cell in row:
                if cell.value:
                    lines = str(cell.value).count("\n") + 1
                    # Base height 15, slightly more for breathing room
                    required_height = lines * 16 
                    max_height = max(max_height, required_height)
            if max_height > 0:
                sheet.row_dimensions[row[0].row].height = max_height


